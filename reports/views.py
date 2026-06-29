from django.shortcuts import render

# Create your views here.
from datetime import date
from calendar import month_name

from django.db.models import Sum
from django.http import HttpResponse

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.views import APIView, Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

from expenses.models import ExpenseModel, BudgetModel


class PDFReportAPIView(APIView):
    
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):

        today = date.today()

        month = int(request.query_params.get("month", today.month))
        year = int(request.query_params.get("year", today.year))
        if month < 1 or month > 12:
            return Response(
                {"error": "Invalid month"},
                    status=400
                )
        month_display = f"{month_name[month]} {year}"

        expenses = ExpenseModel.objects.filter(
            user=request.user,
            date__month=month,
            date__year=year,
        ).order_by("date")

        budget = BudgetModel.objects.filter(
            user=request.user,
            month=month,
            year=year,
        ).first()

        total_spent = (
            expenses.aggregate(
                total=Sum("amount")
            )["total"] or 0
        )

        budget_amount = budget.budget if budget else 0

        remaining = budget_amount - total_spent

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = (
            f'attachment; filename="Expense_Report_{month}_{year}.pdf"'
        )

        document = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        # Title
        elements.append(
            Paragraph(
                "SmartSpend AI",
                styles["Title"]
            )
        )

        elements.append(
            Paragraph(
                f"Expense Report - {month_display}",
                styles["Heading2"]
            )
        )

        elements.append(Spacer(1, 20))

        # Summary
        elements.append(
            Paragraph(
                f"<b>Budget :</b> ₹{budget_amount}",
                styles["Normal"]
            )
        )

        elements.append(
            Paragraph(
                f"<b>Spent :</b> ₹{total_spent}",
                styles["Normal"]
            )
        )

        elements.append(
            Paragraph(
                f"<b>Remaining :</b> ₹{remaining}",
                styles["Normal"]
            )
        )

        elements.append(Spacer(1, 20))

        # Expense Table
        table_data = [
            [
                "Date",
                "Item",
                "Category",
                "Amount",
            ]
        ]

        for expense in expenses:

            table_data.append(
                [
                    expense.date.strftime("%d-%m-%Y"),
                    expense.item,
                    expense.category,
                    f"₹{expense.amount}",
                ]
            )

        table = Table(table_data)

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

                    ("GRID", (0, 0), (-1, -1), 1, colors.black),

                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ]
            )
        )

        elements.append(table)

        document.build(elements)

        return response


class ExcelReportAPIView(APIView):

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):

        today = date.today()

        month = int(request.query_params.get("month", today.month))
        year = int(request.query_params.get("year", today.year))
        if month < 1 or month > 12:
            return Response(
                { "error": "Invalid month"},
                    status=400
                )
        month_display = f"{month_name[month]} {year}"

        expenses = ExpenseModel.objects.filter(
            user=request.user,
            date__month=month,
            date__year=year
        ).order_by("date")

        budget = BudgetModel.objects.filter(
            user=request.user,
            month=month,
            year=year
        ).first()

        total_spent = (
            expenses.aggregate(
                total=Sum("amount")
            )["total"] or 0
        )

        budget_amount = budget.budget if budget else 0

        remaining = budget_amount - total_spent

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = f"{month_name[month]} Report"

        # Title
        sheet["A1"] = "SmartSpend AI"
        sheet["A1"].font = Font(size=18, bold=True)

        sheet["A2"] = f"Expense Report - {month_display}"
        sheet["A2"].font = Font(size=14, bold=True)

        # Summary
        sheet["A4"] = "Budget"
        sheet["B4"] = float(budget_amount)

        sheet["A5"] = "Spent"
        sheet["B5"] = float(total_spent)

        sheet["A6"] = "Remaining"
        sheet["B6"] = float(remaining)

        # Table Header
        headers = [
            "Date",
            "Item",
            "Category",
            "Amount",
            "Notes",
            "Recurring",
        ]

        row = 8

        for col, header in enumerate(headers, start=1):

            cell = sheet.cell(row=row, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        # Expense Data
        row = 9

        for expense in expenses:

            sheet.cell(row=row, column=1).value = expense.date.strftime("%d-%m-%Y")
            sheet.cell(row=row, column=2).value = expense.item
            sheet.cell(row=row, column=3).value = expense.category
            sheet.cell(row=row, column=4).value = float(expense.amount)
            sheet.cell(row=row, column=5).value = expense.notes
            sheet.cell(row=row, column=6).value = "Yes" if expense.is_recurring else "No"

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response[
            "Content-Disposition"
        ] = f'attachment; filename="Expense_Report_{month}_{year}.xlsx"'

        workbook.save(response)

        return response